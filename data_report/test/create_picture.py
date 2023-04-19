import xlsxwriter
import openpyxl
from openpyxl import load_workbook
import get_data
import xlrd
import matplotlib as mpl
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd

workbook = xlsxwriter.Workbook('pic.xlsx')
worksheet = workbook.add_worksheet()
bold = workbook.add_format({'bold': 1})
headings = ['状态', '数据']
worksheet.write_row('A1', headings, bold)
worksheet.write_column('A2',get_data.data1)
worksheet.write_column('B2',get_data.data2)
chart = workbook.add_chart({'type': 'column'})
chart.add_series({
        'categories': '=Sheet1!$A$2:$A$10',
        'values': '=Sheet1!$B$2:$B$10',
        'line': {'color':'red'},
})
chart.set_title({'name':'缺陷解决情况'})
chart.set_x_axis({'name': '缺陷解决状态'})
chart.set_y_axis({'name': '各解决状态对应数量'})
worksheet.insert_chart('C10',chart, {'x_offset': 25, 'y_offset': 10})
workbook.close()


#     sheet1.write(i,1,get_data.ws.cell(i,16).value)





