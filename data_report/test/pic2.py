import xlsxwriter
import openpyxl
from openpyxl import load_workbook
import get_data
from datetime import datetime,timedelta
from collections import Counter

#缺陷解决周期分布图

data = []
for i in range(2,get_data.bug_max):
    if get_data.ws.cell(i,15).value != None:
        data.append(int(get_data.ws.cell(i,15).value) - int(get_data.ws.cell(i,3).value))
    else:
        i = i + 1
result = dict(Counter(data))
# print([key for key,value in result.items()])
# print({key: value for key,value in result.items()})
data1 = list(result.keys())
data2 = list(result.values())


workbook = xlsxwriter.Workbook('pic2.xlsx')
worksheet = workbook.add_worksheet()
bold = workbook.add_format({'bold': 1})
headings = ['缺陷解决周期', '数据']
worksheet.write_row('A1', headings, bold)
worksheet.write_column('A2',data1)
worksheet.write_column('B2',data2)

chart = workbook.add_chart({'type': 'pie'})
chart.add_series({
        'categories': '=Sheet1!$A$2:$A$10',
        'values': '=Sheet1!$B$2:$B$10',
        'line': {'color': 'white'},
})
chart.set_title({'name':'缺陷解决周期分布图'})
worksheet.insert_chart('C10',chart, {'x_offset': 25, 'y_offset': 10})
workbook.close()