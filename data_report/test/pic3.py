import xlsxwriter
import openpyxl
from openpyxl import load_workbook
import get_data
from datetime import datetime,timedelta
from collections import Counter
import pandas as pd


 #输入轮次日期信息
data = []
for i in range(2, get_data.bug_max):
    if get_data.ws.cell(i, 16).value != None:
        data.append(get_data.ws.cell(i,16).value)
    else:
        data = data
# last_bug = 0
# for i in range(2, get_data.bug_max):
#     if data[-2]<get_data.ws.cell(i, 3).value<=data[-1]:
#         last_bug = last_bug + 1
l = len(data)
bug_list_creat = []   #缺陷提交轮次统计
bug_list_close = []   #缺陷关闭轮次统计
bug_list_reopen = []  #缺陷重开轮次统计

bug_creat = []   #用于存放缺陷创建日期
bug_close = []   #用于存放缺陷关闭日期
bug_reopen = []  #用于存放缺陷重开日期

bug_list_creat_total = []  #用于存放轮次累计提交缺陷
bug_list_close_total = []  #用于存放轮次累计关闭缺陷

#   处理缺陷创建轮次
for i in range(2,get_data.bug_max):  #获取缺陷创建日期并存入 bug_creat中
    bug_creat.append(int(get_data.ws.cell(i, 3).value))
for i in range(len(data)):  #初始化缺陷创建统计日期轮次list
    bug_list_creat.append(0)

for j in range(len(bug_creat)):
    if data[-1] <= bug_creat[j]:
        bug_list_creat[-1] = bug_list_creat[-1] + 1
    else:
        for i in range(l):
            if data[i] == bug_creat[j]:
                bug_list_creat[i] = bug_list_creat[i] + 1

#   处理缺陷关闭轮次
for i in range(2,get_data.bug_max):  #获取缺陷关闭日期并存入 bug_close中
    if get_data.ws.cell(i, 6).value != None:
        bug_close.append(int(get_data.ws.cell(i, 6).value))
for i in range(len(data)):  #初始化缺陷关闭统计日期轮次list
    bug_list_close.append(0)

for j in range(len(bug_close)):
    if data[-1] <= bug_creat[j]:
        bug_list_close[-1] = bug_list_close[-1] + 1
    else:
        for i in range(l):
            if data[i] == bug_close[j]:
                bug_list_close[i] = bug_list_close[i] + 1



workbook = xlsxwriter.Workbook('pic3.xlsx')
worksheet = workbook.add_worksheet()
bold = workbook.add_format({'bold': 1})
headings = ['轮次日期', '创建缺陷', '关闭缺陷', '累计提交缺陷数', '累计关闭缺陷数']

#写excel表头   ----各轮新增&关闭故障对比图
worksheet.write_row('A1', headings, bold)
lunci = []
for i in range(1,len(data)+1):
    lunci.append('第' + str(i) + '轮')
list = [lunci, bug_list_creat, bug_list_close]
worksheet.write_column('A2', list[0])
worksheet.write_column('B2', list[1])
worksheet.write_column('C2', list[2])
chart_1 = workbook.add_chart({'type':'column'})
chart_1.add_series({
        'name': '=Sheet1!$B$1',
        'categories': '=Sheet1!$A$2:$A$12',
        'values': '=Sheet1!$B$2:$B$12',
        'line': {'color': 'yellow'},
})
chart_1.add_series({
        'name': '=Sheet1!$C$1',
        'categories': '=Sheet1!$A$2:$A$10',
        'values': '=Sheet1!$C$2:$C$12',
        'line': {'color': 'blue'},
})
chart_1.set_title({'name':'各轮新增&关闭故障对比图'})
chart_1.set_x_axis({'name':'轮次'})
chart_1.set_y_axis({'name':'缺陷数量'})
worksheet.insert_chart('F1',chart_1, {'x_offset': 25, 'y_offset': 10})

#轮次累计提交缺陷和关闭缺陷数量

for i in range(len(bug_list_creat)):
    bug_list_creat_total.append(0)
    bug_list_close_total.append(0)
for i in range(len(bug_list_creat)):
    if i == 0:
        bug_list_creat_total[i] = bug_list_creat[i]+bug_list_creat_total[i]
    else:
        bug_list_creat_total[i] = bug_list_creat[i] + bug_list_creat_total[i-1]
for i in range(len(bug_list_close)):
    if i == 0:
        bug_list_close_total[i] = bug_list_close[i]+bug_list_close_total[i]
    else:
        bug_list_close_total[i] = bug_list_close[i] + bug_list_close_total[i-1]

worksheet.write_column('D2', bug_list_creat_total)
worksheet.write_column('E2', bug_list_close_total)
chart_1 = workbook.add_chart({'type':'line'})
chart_1.add_series({
        'name': '=Sheet1!$D$1',
        'categories': '=Sheet1!$A$2:$A$12',
        'values': '=Sheet1!$D$2:$D$12',
        'line': {'color': 'red'},
})
chart_1.add_series({
        'name': '=Sheet1!$E$1',
        'categories': '=Sheet1!$A$2:$A$12',
        'values': '=Sheet1!$E$2:$E$12',
        'line': {'color': 'blue'},
})
chart_1.set_title({'name':'各轮累计新增&累计关闭故障对比图'})
chart_1.set_x_axis({'name':'轮次'})
chart_1.set_y_axis({'name':'缺陷数量'})
worksheet.insert_chart('N1',chart_1, {'x_offset': 25, 'y_offset': 20})







workbook.close()



