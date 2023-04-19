import xlrd
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime
from collections import Counter


wb = load_workbook('data.xlsx')
ws = wb.active
# sheet.cell(row= , column=,).value， 获取指定cell的值
bug_max = ws.max_row-1
dianbo = 0
dianbo_Highest = 0
dianbo_High = 0
dianbo_Medium = 0
dianbo_low = 0



# 获取第一行指定列的具体列数值
def column_number(str):
    for i in range(1, ws.max_column):
        if ws.cell(row=1, column=i).value == str:
            return i

# 获取对应模块及对应缺陷等级的缺陷数量
# def bug_number(str1,str2=None):
#     bg = 0
#     if str2 == None:
#         for i in range(1, ws.max_row):
#             if ws.cell(row=i, column=8).value == str1:
#                 bg = bg + 1
#         return bg
#     else:
#         for i in range(1, ws.max_row):
#             if ws.cell(row= i ,column=8).value == str1 and ws.cell(row = i, column=2).value == str2:
#                 bg = bg + 1
#         return bg
# 获取对应模块，对应优先级下的缺陷数量
def bug_number(str1,str2,str3,str4=None):
    bg = 0
    if str4 == None:
        for i in range(1, ws.max_row):
            if ws.cell(row=i, column=column_number(str2)).value == str1:
                bg = bg + 1
        return bg
    else:
        for i in range(1, ws.max_row):
            if ws.cell(row= i ,column=column_number(str2)).value == str1 and ws.cell(row = i, column=column_number(str3)).value == str4:
                bg = bg + 1
        return bg


# 获取各解决结果对应的缺陷数量
def bug_conclusion(str1,str2):
    bc = 0
    for i in range(1, ws.max_row):
        if ws.cell(row=i ,column=column_number(str1)).value == str2:
            bc = bc + 1
    return bc




# 获取各缺陷的解决周期
# for i in range(2, ws.max_row):
#     ws.cell(row = i, column = ws.max_column + 1).value = ord(str(ws.cell(row = i, column= 15).value).split('/')[0]) -
#     ord(str(ws.cell(row = i, column= 3).value).split('/')[0])
#     print(ws.cell(row = i, column = ws.max_column + 1).value)

#获取创建图表的数据



