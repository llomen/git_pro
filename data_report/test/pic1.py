import xlsxwriter
import openpyxl
from openpyxl import load_workbook
import get_data
import xlrd
import matplotlib as mpl
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from collections import Counter


# 故障状态分布图

data1 = []
data2 = []
for i in range(2,get_data.bug_max):
    data1.append(get_data.ws.cell(i,11).value)

result = dict(Counter(data1))
# print([key for key,value in result.items()])
# print({key: value for key,value in result.items()})
data1 = list(result.keys())
data2 = list(result.values())

workbook = xlsxwriter.Workbook('pic1.xlsx')
worksheet = workbook.add_worksheet()
bold = workbook.add_format({'bold': 1})
headings = ['状态', '数据']
worksheet.write_row('A1', headings, bold)
worksheet.write_column('A2',data1)
worksheet.write_column('B2',data2)
chart = workbook.add_chart({'type': 'pie'})
chart.add_series({
        'categories': '=Sheet1!$A$2:$A$10',
        'values': '=Sheet1!$B$2:$B$10',
        'line': {'color':'white'},
})
chart.set_title({'name':'故障状态分布图'})
worksheet.insert_chart('C10',chart, {'x_offset': 25, 'y_offset': 10})
workbook.close()
