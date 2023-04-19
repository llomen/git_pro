import xlsxwriter
import openpyxl
from openpyxl import load_workbook
import get_data
import xlrd
import matplotlib as mpl
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from collections import Counter


# 故障状态分布图

data1 = []
data2 = []
for i in range(2, get_data.bug_max+2):
    data1.append(get_data.ws.cell(i,11).value)

print(data1)
result = dict(Counter(data1))
# print([key for key,value in result.items()])
# print({key: value for key,value in result.items()})
data1 = list(result.keys())
data2 = list(result.values())

print(data1,data2)
workbook = xlsxwriter.Workbook('pic1.xlsx')
worksheet = workbook.add_worksheet()
bold = workbook.add_format({'bold': 1})
headings = ['状态', '数据']
worksheet.write_row('A1', headings, bold)
worksheet.write_column('A2',data1)
worksheet.write_column('B2',data2)
chart = workbook.add_chart({'type': 'pie'})
s1 = 'Sheet1!$A$2:$A$' + str(len(data1)+1)
s2 = 'Sheet1!$B$2:$B$' + str(len(data2)+1)
chart.add_series({
        'categories': s1,
        'values': s2,
        'line': {'color':'white'},
})
chart.set_title({'name':'故障状态分布图'})
worksheet.insert_chart('C10',chart, {'x_offset': 25, 'y_offset': 10})
workbook.close()
