#coding=utf-8
import xlsxwriter
import openpyxl
from openpyxl import load_workbook
import get_data
import xlrd
import matplotlib as mpl
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from collections import Counter


# 故障状态分布图
data1 = []
data2 = []
for i in range(2,get_data.bug_max+2):
    data1.append(get_data.ws.cell(i,11).value)
result = dict(Counter(data1))
# print([key for key,value in result.items()])
# print({key: value for key,value in result.items()})
data1 = list(result.keys())
data2 = list(result.values())
excel_name = input()
workbook = xlsxwriter.Workbook(excel_name + '_1.xlsx')
worksheet = workbook.add_worksheet()
bold = workbook.add_format({'bold': 1})
headings = ['状态', '数据']
worksheet.write_row('A1', headings, bold)
len_pie1 = 'Sheet1!$A$2:$A$' + str(len(data1)+1)
len_pie2 = 'Sheet1!$B$2:$B$' + str(len(data2)+1)
worksheet.write_column('A2',data1)
worksheet.write_column('B2',data2)
chart = workbook.add_chart({'type': 'pie'})
chart.add_series({
        'categories': len_pie1,
        'values': len_pie2,
        'line': {'color':'white'},
})
chart.set_title({'name':'故障状态分布图'})
worksheet.insert_chart('C10',chart, {'x_offset': 25, 'y_offset': 10})
workbook.close()

# 缺陷解决周期分布图
data = []
for i in range(2,get_data.bug_max+2):
    if get_data.ws.cell(i,15).value != None:
        data.append(int(get_data.ws.cell(i,15).value) - int(get_data.ws.cell(i,3).value))
    else:
        i = i + 1
result = dict(Counter(data))
# print([key for key,value in result.items()])
# print({key: value for key,value in result.items()})
data3 = list(result.keys())
data4 = list(result.values())
workbook = xlsxwriter.Workbook(excel_name + '_2.xlsx')
worksheet = workbook.add_worksheet()
bold = workbook.add_format({'bold': 1})
headings = ['缺陷解决周期', '数据']
worksheet.write_row('A1', headings, bold)
worksheet.write_column('A2', data3)
worksheet.write_column('B2', data4)
lne_pie3 = 'Sheet1!$A$2:$A$' + str(len(data3)+1)
lne_pie4 = 'Sheet1!$B$2:$B$' + str(len(data4)+1)
chart = workbook.add_chart({'type': 'pie'})
chart.add_series({
        'categories': lne_pie3,
        'values': lne_pie4,
        'line': {'color': 'white'},
})
chart.set_title({'name':'缺陷解决周期分布图'})
worksheet.insert_chart('C10',chart, {'x_offset': 25, 'y_offset': 10})
workbook.close()



data = []
for i in range(2, get_data.bug_max+2):
    if get_data.ws.cell(i, 16).value != None:
        data.append(get_data.ws.cell(i,16).value)
    else:
        data = data
# last_bug = 0
# for i in range(2, get_data.bug_max+2):
#     if data[-2]<get_data.ws.cell(i, 3).value<=data[-1]:
#         last_bug = last_bug + 1
l = len(data)
bug_list_creat = []   #缺陷提交轮次统计
bug_list_close = []   #缺陷关闭轮次统计
bug_list_reopen = []  #缺陷重开轮次统计

bug_creat = []   #用于存放缺陷创建日期
bug_close = []   #用于存放缺陷关闭日期
bug_reopen = []  #用于存放缺陷重开日期

bug_list_creat_total = []  #用于存放轮次累计提交缺陷
bug_list_close_total = []  #用于存放轮次累计关闭缺陷

for i in range(2,get_data.bug_max+2):  #获取缺陷创建日期并存入 bug_creat中
        bug_creat.append(get_data.ws.cell(i, 3).value)
# for i in range(len(data)):  #初始化缺陷创建统计日期轮次list
#     bug_list_creat.append(0)
for i in data:
    bug_creat.append(i)
bug_creat.sort()
mm1 = []
mm3 = []
for j in range(len(bug_creat)):
    for k in range(l):
        if data[k] == bug_creat[j]:
            mm1.append(j)
for i in range(1,len(mm1)):
    bug_list_creat.append(mm1[i]-mm1[i-1]-1)
bug_list_creat.append(len(bug_creat)-mm1[-1]-1)



for i in range(2,get_data.bug_max+2):  #获取缺陷关闭日期并存入 bug_close中
    if get_data.ws.cell(i, 15).value != None and get_data.ws.cell(i, 11).value != 'Unresolved':
        bug_close.append(get_data.ws.cell(i, 15).value)
for i in data:
    bug_close.append(i)
bug_close.sort()

for j in range(len(bug_close)):
    for k in range(l):
        if data[k] == bug_close[j]:
            mm3.append(j)
for i in range(1,len(mm3)):
    bug_list_close.append(mm3[i]-mm3[i-1]-1)
bug_list_close.append(len(bug_close)-mm3[-1]-1)


workbook = xlsxwriter.Workbook(excel_name + '_3.xlsx')
worksheet = workbook.add_worksheet()
bold = workbook.add_format({'bold': 1})
headings = ['轮次日期', '创建缺陷', '关闭缺陷', '累计提交缺陷数', '累计关闭缺陷数']

#写excel表头   ----各轮新增&关闭故障对比图
worksheet.write_row('A1', headings, bold)
lunci = []
for i in range(1,len(data)+1):
    lunci.append('第' + str(i) + '轮')
list = [lunci, bug_list_creat, bug_list_close]
worksheet.write_column('A2', list[0])
worksheet.write_column('B2', list[1])
worksheet.write_column('C2', list[2])
len_column1 = '=Sheet1!$A$2:$A$' + str(len(list[0])+1)
len_column2 = '=Sheet1!$B$2:$B$' + str(len(list[1])+1)
len_column3 = '=Sheet1!$C$2:$C$' + str(len(list[2])+1)

chart_1 = workbook.add_chart({'type':'column'})
chart_1.add_series({
        'name': '=Sheet1!$B$1',
        'categories': len_column1,
        'values': len_column2,
        'line': {'color': 'yellow'}
})
chart_1.add_series({
        'name': '=Sheet1!$C$1',
        'categories': len_column1,
        'values': len_column3,
        'line': {'color': 'blue'},
})
chart_1.set_title({'name':'各轮新增&关闭故障对比图'})
chart_1.set_x_axis({'name':'轮次'})
chart_1.set_y_axis({'name':'缺陷数量'})
worksheet.insert_chart('F1',chart_1, {'x_offset': 25, 'y_offset': 10})

#轮次累计提交缺陷和关闭缺陷数量

for i in range(len(bug_list_creat)):
    bug_list_creat_total.append(0)
    bug_list_close_total.append(0)
for i in range(len(bug_list_creat)):
    if i == 0:
        bug_list_creat_total[i] = bug_list_creat[i]+bug_list_creat_total[i]
    else:
        bug_list_creat_total[i] = bug_list_creat[i] + bug_list_creat_total[i-1]
for i in range(len(bug_list_close)):
    if i == 0:
        bug_list_close_total[i] = bug_list_close[i]+bug_list_close_total[i]
    else:
        bug_list_close_total[i] = bug_list_close[i] + bug_list_close_total[i-1]

worksheet.write_column('D2', bug_list_creat_total)
worksheet.write_column('E2', bug_list_close_total)
len_line1 = '=Sheet1!$D$2:$D$' + str(len(bug_list_creat_total)+1)
len_line2 = '=Sheet1!$E$2:$E$' + str(len(bug_list_close_total)+1)
chart_1 = workbook.add_chart({'type':'line'})
chart_1.add_series({
        'name': '=Sheet1!$D$1',
        'categories': len_column1,
        'values': len_line1,
        'line': {'color': 'red'},
})
chart_1.add_series({
        'name': '=Sheet1!$E$1',
        'categories': len_column1,
        'values': len_line2,
        'line': {'color': 'blue'},
})
chart_1.set_title({'name':'各轮累计新增&累计关闭故障对比图'})
chart_1.set_x_axis({'name':'轮次'})
chart_1.set_y_axis({'name':'缺陷数量'})
worksheet.insert_chart('N1',chart_1, {'x_offset': 25, 'y_offset': 20})
workbook.close()